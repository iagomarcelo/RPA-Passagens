from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
import time
import os


def openWorkbook(archiveName):  
    templatePath = os.getcwd()+"\\"+ archiveName +".xlsx"
    workbook = load_workbook(templatePath)
    workbookPage = workbook.active
    return workbookPage

# 'Dinheiro' : '', 
def gettingExcelData(workbookPage = openWorkbook(archiveName='template')):
    sitesDict = {'Latam Pass' : 'https://latampass.latam.com/', 'Smiles' : 'https://www.smiles.com.br/', 'Tudo Azul' : 'https://interline.tudoazul.com/', 'MaxMilhas' : 'https://www.maxmilhas.com.br/', 'Skyscanner' : 'https://www.skyscanner.com.br/', 'Voe Livre' : 'https://www.voelivre.com.br/', 'Decolar' : 'https://www.decolar.com/', 'Hotmilhas': 'https://hotmilhas.com.br/', 'Tap Miles & Go' : 'https://www.flytap.com/pt-br/miles-and-go', 'Jetsmart' : 'https://jetsmart.com/', 'Azul Interlines' : 'https://www.voeazul.com.br/', 'Iberia' : 'https://www.iberia.com/', 'Viajanet' : 'https://www.viajanet.com.br/', '123 Milhas' : 'https://123milhas.com/'}



    origem = workbookPage.cell(row = 6, column = 4).value
    destino = workbookPage.cell(row = 11, column = 4).value
    idaVolta = workbookPage.cell(row = 5, column = 7).value 
    dataIda = str(workbookPage.cell(row = 8, column = 7).value) + '/' + str(workbookPage.cell(row = 8, column = 9).value) + '/' + str(workbookPage.cell(row = 8, column = 11).value)
    dataVolta = str(workbookPage.cell(row = 11, column = 7).value) + '/' + str(workbookPage.cell(row = 11, column = 9).value) + '/' + str(workbookPage.cell(row = 11, column = 11).value)
    qtdAdultos = workbookPage.cell(row = 6, column = 14).value
    qtdCriancas = workbookPage.cell(row = 7, column = 14).value 
    qtdBebes = workbookPage.cell(row = 8, column = 14).value
    classe = workbookPage.cell(row = 1, column = 14).value

    #print(origem, destino, idaVolta, dataIda, dataVolta, qtdAdultos, qtdCriancas, qtdBebes, classe)
    return sitesDict
    
    

def openBrowser(url):
    browser = webdriver.Chrome()
    browser.get(url=url)
    browser.maximize_window()
    time.sleep(3)

