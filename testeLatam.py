from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
import time
import os

templatePath = os.getcwd()+"\\"+ 'template' +".xlsx"
workbook = load_workbook(templatePath)
workbookPage = workbook.active

origem = workbookPage.cell(row = 6, column = 4).value
destino = workbookPage.cell(row = 11, column = 4).value
idaVolta = workbookPage.cell(row = 5, column = 7).value 

diaIda = str(workbookPage.cell(row = 8, column = 7).value)
mesIda = str(workbookPage.cell(row = 8, column = 9).value)
diaVolta = str(workbookPage.cell(row = 11, column = 7).value)
mesVolta = str(workbookPage.cell(row = 11, column = 9).value)

if len(diaIda) == 1:
    diaIda = '0' + diaIda

if len(mesIda) == 1:
    mesIda = '0' + mesIda
    
if len(diaVolta) == 1:
    diaVolta = '0' + diaVolta

if len(mesVolta) == 1:
    mesVolta = '0' + mesVolta

qtdAdultos = workbookPage.cell(row = 6, column = 14).value
qtdCriancas = workbookPage.cell(row = 7, column = 14).value 
qtdBebes = workbookPage.cell(row = 8, column = 14).value
classe = workbookPage.cell(row = 1, column = 14).value

# print ({'origem': origem, 'destino': destino, 'idaVolta': idaVolta, 'dataIda': dataIda, 'dataVolta': 'dataVolta', 'qtdAdultos': qtdAdultos, 'qtdCriancas': qtdCriancas, 'qtdBebes': qtdBebes, 'classe': classe})